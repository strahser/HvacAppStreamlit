from importlib.resources import path
import pandas as pd
import xlwings as xw
import os
import numpy as np
import json
from library_hvac_app.list import *
import io

class Loader:
    def __init__(self, data_path,) -> None:
        self.data_path = data_path

    def load_json(self):
        with open(self.data_path, "r") as read_file:
            data = json.load(read_file)
        return data


    def load_json_pd(self, idx='S_ID'):
        
        if isinstance(self.data_path,pd.DataFrame):
            return self.data_path
        elif isinstance(self.data_path,dict):
            df_out = pd.DataFrame.from_dict(self.data_path)
            df = df_out.T
            df = df.rename_axis(idx).reset_index()
            df[idx] = df[idx].astype(str)
            return df
        elif os.path.splitext(self.data_path)[1] == ".json":
            df_out = pd.read_json(self.data_path)
            df = df_out.T
            df = df.rename_axis(idx).reset_index()
            df[idx] = df[idx].astype(str)
            return df
        else:
            df = self.data_path
            return df

def data_frame_to_excel_open(df, filename, sheet_name, start_position="A1"):
    """
    open excel file and add df
    """
    app = xw.App(visible=False)
    book = app.books.open(filename)
    ws = book.sheets[sheet_name]
    ws.clear()
    ws.range(start_position).options(pd.DataFrame, index=False).value = df
    table = ws.tables.add(source=ws[start_position].expand())
    ws.autofit("c")
    book.save()
    book.close()
    app.quit()

def group_subgroups(df, key_group_index, columns_list, agg_column, total_name):
    """
    groping and make pivot table for example all heat loss for room, (surfaces,window,doors)
    df:pd.df with column for agg,
    key_group_index => room name 
    columns_list => none agg columns area,surfaces name
    agg_column => heat load of surface for agg
    total_name => str, Total for example
    subgroups = group_subgroups(
                        df=config.df_merge_surf,
                        key_group_index= "S_ID",
                        columns_list=[
                            'surf_area', 'surf_name',
                            'orientation', 'heat_sun_key',
                            'k-ef', 't_out', 'k-orientation'
                            ],
                        agg_column= 'heat_load, w/m2',
                        total_name="Total"
    """
    group = df.groupby([key_group_index]).\
        apply(lambda sub_df: sub_df.
              pivot_table(index=columns_list, values=[agg_column],
                          aggfunc=np.sum, margins=True, margins_name=total_name
                          ))
    return group


def result_to_excel_add_table(df_dict: dict, excel_book_path: path, index=False, convert_to_table=True,add_format =False):
    """
    create new file excel. convert_to_table == True -> make smarte excel table.
    df_dict {sheet_name:df}
    excel_book_path path to excel
    """
    writer = pd.ExcelWriter(excel_book_path, engine='xlsxwriter')
    for sh_name in df_dict.keys():
        if convert_to_table == True:
            df_dict[sh_name].to_excel(
                writer,
                sheet_name=sh_name,
                startrow=1, startcol=0,
                header=False,
                index=index,
                freeze_panes=(1, 1),
                float_format="%.2f")

            workbook = writer.book
            worksheet = writer.sheets[sh_name]
            column_settings = [{'header': column}
                               for column in df_dict[sh_name]]
            (max_row, max_col) = df_dict[sh_name].shape
            worksheet.add_table(0, 0, max_row, max_col - 1,
                                {'columns': column_settings,
                                 'banded_columns': True,
                                 'name': sh_name,
                                 'style': 'Table Style Light 8'})
        if add_format:
            worksheet.set_landscape()
            worksheet.set_paper(9)
            worksheet.set_margins(left=1,right=0.2,top =0.2,bottom =0.2 )
            worksheet.set_header('&C&"Courier New,Bold Italic"Test_name')
            worksheet.set_footer('&CPage &P of &N')
            worksheet.repeat_rows(0)
            # https://fooobar.com/questions/16142772/painting-a-cell-in-excel-with-condition-using-python
            # https://xlsxwriter.readthedocs.io/worksheet.html#conditional_format%23conditional_format
            # https://stackoverflow.com/questions/44150078/python-using-pandas-to-format-excel-cell
            # Add a format for pass. Green fill with dark green text.
            pass_format = workbook.add_format({'bg_color': '#C6EFCE',
                                               'font_color': '#006100'})
            worksheet.conditional_format(0, 0, max_row, max_col, {'type': 'text',
                                                                  'criteria': 'containing',
                                                                  'value': "define value separatly",  # todo
                                                                  'format': pass_format})
            # worksheet.set_column(0, max_col - 1, 25)

        else:
            df_dict[sh_name].to_excel(writer, sheet_name=sh_name,
                                      startrow=0, startcol=0, header=True, index=index, freeze_panes=(1, 1))
            workbook = writer.book
            worksheet = writer.sheets[sh_name]
    writer.save()


def number_to_float_format(value):
    if isinstance(value, float):
        res = '{:,.0f}'.format(value).replace(
            ',', ' ') if value > 1e3 else '{:,.2f}'.format(value)
    elif isinstance(value, int):
        res = '{:,.0f}'.format(value).replace(
            ',', ' ') if value > 1e3 else '{:,.0f}'.format(value)
    else:
        res = value
    return res


def df_rows_to_float_format(df_rows):
    res = []
    for lst in df_rows:
        float_form = [number_to_float_format(sub_list) for sub_list in lst]
        res.append(float_form)
    return res


def df_html_format(df, table_id="t1", index=True, escape=False):

    res = df.to_html(classes=[" table-hover table-striped table-light\
    table  text-center table-condensed  thead-dark table-bordered border-info "],

                     na_rep='-',
                     index=index,
                     table_id=table_id,
                     escape=escape,
                     float_format=lambda x: '{:,.0f}'.format(x).replace(u',', u' ') if x > 1e3 else '{:,.2f}'.format(x))
    res = res.replace("\n", "").replace(
        "(", "").replace(")", "").replace("'", "")
    return res


def data_frame_to_excel_open(
        df: pd.DataFrame,
        filename: os.path,
        sheet_name: str,
        start_position: str = "A1"):
    """ open excel book.choose sheet. choose start row column. save df to sheet
    based on xw!.

    Args:
        df (pd.DataFrame): df to save
        filename (os.path): path to file
        sheet_name (str): excel sheet name
        start_position (str, optional):  Defaults to "A1".
    """
    app = xw.App(visible=False)
    book = app.books.open(filename)
    ws = book.sheets[sheet_name]
    ws.clear()
    ws.range(start_position).options(pd.DataFrame, index=False).value = df
    ws.tables.add(source=ws[start_position].expand())
    ws.autofit("c")
    book.save()
    book.close()
    app.quit()


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, startcol=None,
    truncate_sheet=False, resizeColumns=True, na_rep = 'NA', **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file

      resizeColumns: default = True . It resize all columns based on cell content width
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]
      na_rep: default = 'NA'. If, instead of NaN, you want blank cells, just edit as follows: na_rep=''


    Returns: None

    *******************

    CONTRIBUTION:
    Current helper function generated by [Baggio]: https://stackoverflow.com/users/14302009/baggio?tab=profile
    Contributions to the current helper function: https://stackoverflow.com/users/4046632/buran?tab=profile
    Original helper function: (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)


    Features of the new helper function:
    1) Now it works with python 3.9 and latest versions of pandas and openpxl
    ---> Fixed the error: "zipfile.BadZipFile: File is not a zip file".
    2) Now It resize all columns based on cell content width AND all variables will be visible (SEE "resizeColumns")
    3) You can handle NaN,  if you want that NaN are displayed as NaN or as empty cells (SEE "na_rep")
    4) Added "startcol", you can decide to start to write from specific column, oterwise will start from col = 0

    *******************



    """
    from openpyxl import load_workbook
    from string import ascii_uppercase
    from openpyxl.utils import get_column_letter
    from openpyxl import Workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    try:
        f = open(filename)
        # Do something with the file
    except IOError:
        # print("File not accessible")
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        wb.save(filename)
        wb.close()

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')


    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        # startrow = -1
        startrow = 0

    if startcol is None:
        startcol = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, startcol=startcol, na_rep=na_rep, **to_excel_kwargs)


    if resizeColumns:

        ws = writer.book[sheet_name]

        def auto_format_cell_width(ws):
            for letter in range(1,ws.max_column):
                maximum_value = 0
                for cell in ws[get_column_letter(letter)]:
                    val_to_check = len(str(cell.value))
                    if val_to_check > maximum_value:
                        maximum_value = val_to_check
                ws.column_dimensions[get_column_letter(letter)].width = maximum_value + 2

        auto_format_cell_width(ws)

    # save the workbook
    writer.save()

def df_to_excel_in_memory(df_list,sheet_list,is_index_need =False):
    df_list = to_list(df_list)
    sheet_list = to_list(sheet_list)
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
        # Write each dataframe to a different worksheet.
        for df_,shet_name in zip(df_list,sheet_list):
            df_.to_excel(writer, sheet_name=shet_name,index=is_index_need)

        writer.save()
    return buffer



def add_column_format(df_,writer,shet_name):
        workbook = writer.book
        worksheet = writer.sheets[shet_name]
        max_row, max_col = df_.shape
        red_format = workbook.add_format({'bg_color': 'red'})
        yellow_format = workbook.add_format({'bg_color': 'yellow'})
        key_column = df_.columns.get_loc("k_ef")
        worksheet.conditional_format(1,key_column,max_row, key_column, {'type':'cell',
                                            'criteria': '>=',
                                            'value': 0.9,  # todo
                                            'format': red_format})
        worksheet.conditional_format(1,key_column,max_row, key_column, {'type':'cell',
                                            'criteria': '<=',
                                            'value': 0.7,  # todo
                                            'format': yellow_format})
def get_col_widths(dataframe):
    # First we find the maximum length of the index column
    idx_max = max([len(str(s)) for s in dataframe.index.values] +
                  [len(str(dataframe.index.name))])
    # Then, we concatenate this to the max of the lengths of column name and its values for each column, left to right
    return [idx_max] + [max([len(str(s)) for s in dataframe[col].values] + [len(col)]) for col in dataframe.columns]

